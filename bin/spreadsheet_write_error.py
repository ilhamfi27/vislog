import sys
import json
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_file_path = sys.argv[1]
output_file_path = sys.argv[2]
errors_json = sys.stdin.readline()

errors = json.loads(errors_json)
# errors = [{'row': 2, 'errors': ['Duplicate SKU']}]

wb = load_workbook(input_file_path)
ws = wb.active
max_column = ws.max_column
while ws.cell(row = 1, column = max_column).value == None:
    max_column -= 1
    
neighbor_cell = ws.cell(row = 1, column = max_column - 1)
cell = ws.cell(row = 1, column = max_column + 1)
cell.fill = neighbor_cell.fill
cell.font = neighbor_cell.font
cell.border = neighbor_cell.border
cell.value = 'Error'

for error in errors:
    row = error['row']
    message = "\n".join(error['errors'])
    cell = ws.cell(row = row, column = max_column + 1)
    cell.value = message

    row_dim = ws.row_dimensions[row]
    row_dim.fill = PatternFill('solid', 'd50000', 'd50000')

wb.save(output_file_path)